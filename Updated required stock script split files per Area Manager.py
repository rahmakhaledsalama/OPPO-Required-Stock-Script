
import pandas as pd
import os
import re
from datetime import date
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


csv_path = r"D:\required stock\copied files to share with area managers\Required Stock Follow-up - All Models Single Sheet - Cairo 2 - Customized Template - Updated @ 04-06-2025(1).csv" # **Update this path**
df = pd.read_csv(csv_path)


# Drop the combined column
df.drop('area manager and model', axis=1, errors='ignore', inplace=True)

# === Output Directory ===
output_dir = r"D:\required stock\split_output"  ## just for test 
os.makedirs(output_dir, exist_ok=True)
today = date.today().strftime("%d-%m-%Y")

# === Format Excel Tabs ===
def format_sheet(sheet):
    for col_idx, col in enumerate(sheet.columns, 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for row_idx, cell in enumerate(col, 1):
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )
            if row_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="ffff00", end_color="ffff00", fill_type="solid")
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[col_letter].width = max_length + 3

    # Remove last 2 columns if they are ["difference [+/-]", etc.]
    if sheet.max_column >= 11:
        sheet.delete_cols(10, 2)

# === Save each Area Manager’s data ===
for full_name, group in df.groupby("area manager"):
    if pd.isna(full_name):
        continue  # skip NaN

    safe_name = re.sub(r'[\\/*?:"<>|]', "", full_name).strip()
    file_name = f"{safe_name} Required stock {today}.xlsx"
    file_path = os.path.join(output_dir, file_name)

    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        for model_name, model_group in group.groupby("model"):
            sheet_name = re.sub(r'[\\/*?:"<>|]', "", model_name).strip()[:31]
            model_group.to_excel(writer, sheet_name=sheet_name, index=False)

    # Format sheets
    wb = load_workbook(file_path)
    for sheet in wb.worksheets:
        format_sheet(sheet)
    wb.save(file_path)

print("✅ Success: Each Area Manager now has a separate Excel file with clean, formatted tabs per model.")